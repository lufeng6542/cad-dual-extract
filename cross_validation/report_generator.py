# -*- coding: utf-8 -*-
"""验证报告生成器 - 4工作表 Excel"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class ValidationReportGenerator:
    def __init__(self):
        self.header_font = Font(bold=True, size=11)
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4",
                                        fill_type="solid")
        self.header_font_white = Font(bold=True, size=11, color="FFFFFF")
        self.green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE",
                                       fill_type="solid")
        self.red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE",
                                     fill_type="solid")
        self.yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C",
                                        fill_type="solid")
        self.border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

    def generate(self, validation: dict, result_a, result_b, output_path: str):
        wb = Workbook()

        self._write_summary(wb, validation, result_a, result_b)
        self._write_matched(wb, validation)
        self._write_discrepancies(wb, validation)
        self._write_missing(wb, validation)

        wb.save(output_path)
        print(f"\n验证报告已保存: {output_path}")

    def _write_summary(self, wb, validation, result_a, result_b):
        ws = wb.active
        ws.title = "汇总统计"
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15

        self._write_header(ws, 1, ["类别", "方案A数量", "方案B数量", "匹配数", "仅A", "仅B"])

        row = 2
        for cat in ["texts", "pipes", "equipment", "dimensions"]:
            s = validation["summary"][cat]
            vals = [
                {"texts": "文字", "pipes": "管道", "equipment": "设备/图块",
                 "dimensions": "标注"}[cat],
                s.get("a_count", 0), s.get("b_count", 0),
                s.get("matched", 0), s.get("a_only", s.get("discrepancies", 0)),
                s.get("b_only", 0)
            ]
            for col, v in enumerate(vals, 1):
                ws.cell(row=row, column=col, value=v).border = self.border

        row += 2
        ws.cell(row=row, column=1, value="耗时").font = self.header_font
        ws.cell(row=row+1, column=1, value="方案A (ezdxf)")
        ws.cell(row=row+1, column=2, value=f"{validation['summary']['time_a']:.2f}s")
        ws.cell(row=row+2, column=1, value="方案B (AI视觉)")
        ws.cell(row=row+2, column=2, value=f"{validation['summary']['time_b']:.2f}s")

        row += 4
        ws.cell(row=row, column=1, value="需人工复核").font = self.header_font
        ws.cell(row=row+1, column=1, value=f"{len(validation.get('review_items', []))} 项")

    def _write_matched(self, wb, validation):
        ws = wb.create_sheet("一致项")
        self._write_header(ws, 1, ["类别", "项目名称", "方案A值", "方案B值",
                                    "偏差%", "匹配类型", "置信度"])

        row = 2
        for m in validation.get("text", {}).get("matched", []):
            self._write_row(ws, row, [
                "文字", m["item_a"].content[:50], m["item_a"].text_type or "",
                m["item_b"].text_type or "", "", m["match_type"],
                m["similarity"]
            ], self.green_fill)
            row += 1

        for m in validation.get("pipe", {}).get("matched", []):
            self._write_row(ws, row, [
                "管道", m["item_a"].system_type, f"{m['item_a'].length}m",
                f"{m['item_b'].length}m", m["deviation_pct"], "长度匹配", "≥0.9"
            ], self.green_fill)
            row += 1

        for m in validation.get("equipment", {}).get("matched", []):
            self._write_row(ws, row, [
                "设备", m["item_a"].equipment_type, m["item_a"].count,
                m["item_b"].count, "", "名称匹配", m["name_similarity"]
            ], self.green_fill)
            row += 1

        for m in validation.get("dimension", {}).get("matched", []):
            self._write_row(ws, row, [
                "标注", m["item_a"].value, m["item_a"].value,
                m["item_b"].value, m.get("deviation_pct", 0), "值匹配", ""
            ], self.green_fill)
            row += 1

    def _write_discrepancies(self, wb, validation):
        ws = wb.create_sheet("差异项")
        self._write_header(ws, 1, ["类别", "项目名称", "方案A值", "方案B值",
                                    "偏差", "状态", "建议"])

        row = 2
        for d in validation.get("pipe", {}).get("discrepancies", []):
            self._write_row(ws, row, [
                "管道", d["item_a"].system_type, f"{d['item_a'].length}m",
                f"{d['item_b'].length}m", f"{d['deviation_pct']}%",
                d["status"], "检查原图纸确认"
            ], self.yellow_fill)
            row += 1

        for d in validation.get("equipment", {}).get("discrepancies", []):
            self._write_row(ws, row, [
                "设备", d["item_a"].equipment_type, d["item_a"].count,
                d["item_b"].count, "", d["status"], "人工核对"
            ], self.yellow_fill)
            row += 1

        for r in validation.get("review_items", []):
            self._write_row(ws, row, [
                r["category"], r["key"], r["value_a"],
                r["value_b"], str(r["deviation"]), "需复核", "建议人工确认"
            ], self.red_fill)
            row += 1

    def _write_missing(self, wb, validation):
        ws = wb.create_sheet("缺失项")
        self._write_header(ws, 1, ["类别", "内容", "来源", "置信度"])

        row = 2
        for t in validation.get("text", {}).get("a_only", []):
            self._write_row(ws, row, ["文字(仅A)", t.content[:60], "ezdxf解析",
                                       t.confidence])
            row += 1
        for t in validation.get("text", {}).get("b_only", []):
            self._write_row(ws, row, ["文字(仅B)", t.content[:60], "AI视觉",
                                       t.confidence])
            row += 1

        for p in validation.get("pipe", {}).get("a_only", []):
            self._write_row(ws, row, ["管道(仅A)", f"{p.system_type}: {p.length}m",
                                       "ezdxf解析", p.confidence])
            row += 1
        for p in validation.get("pipe", {}).get("b_only", []):
            self._write_row(ws, row, ["管道(仅B)", f"{p.system_type}: {p.length}m",
                                       "AI视觉", p.confidence])
            row += 1

        for e in validation.get("equipment", {}).get("a_only", [])[:50]:
            self._write_row(ws, row, ["设备(仅A)", f"{e.equipment_type} x{e.count}",
                                       "ezdxf解析", e.confidence])
            row += 1
        for e in validation.get("equipment", {}).get("b_only", [])[:50]:
            self._write_row(ws, row, ["设备(仅B)", f"{e.equipment_type} x{e.count}",
                                       "AI视觉", e.confidence])
            row += 1

    def _write_header(self, ws, row, headers):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = self.header_font_white
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border
            ws.column_dimensions[get_column_letter(col)].width = 18

    def _write_row(self, ws, row, values, fill=None):
        for col, v in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=v)
            cell.border = self.border
            if fill:
                cell.fill = fill
